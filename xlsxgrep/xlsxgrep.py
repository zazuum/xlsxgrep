#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import argparse
import re
import warnings
import logging
import os
import functools
from concurrent.futures import ProcessPoolExecutor
import pyexcel as p
import xlrd
from openpyxl import load_workbook
from pathlib import Path
import locale
from textwrap import dedent

__license__ = "MIT"
__version__ = "0.0.37"
__author__ = "Ivan Cvitic"
__email__ = "cviticivan@gmail.com"
VERSION_INFO = [
    "xlsxgrep version: {0}".format(__version__),
    "Python version: {0}".format(
        " ".join(line.strip() for line in sys.version.splitlines())
    ),
    "Locale: {0}".format(".".join(str(s) for s in locale.getlocale())),
]

POSIX_CLASSES = {
    "[:alnum:]": r"a-zA-Z0-9",
    "[:alpha:]": r"a-zA-Z",
    "[:blank:]": r" \t",
    "[:cntrl:]": r"\x00-\x1f\x7f",
    "[:digit:]": r"0-9",
    "[:graph:]": r"\x21-\x7e",
    "[:lower:]": r"a-z",
    "[:print:]": r"\x20-\x7e",
    "[:punct:]": r"\x21-\x2f\x3a-\x40\x5b-\x60\x7b-\x7e",
    "[:space:]": r" \t\r\n\v\f",
    "[:upper:]": r"A-Z",
    "[:xdigit:]": r"0-9a-fA-F",
}


def translate_posix_ere(pattern):
    for posix_cls, repl in POSIX_CLASSES.items():
        pattern = pattern.replace(posix_cls, repl)
    return pattern


def check_optional_args(opts, val):
    if opts.get("compiled_regex"):
        return opts["compiled_regex"].search(str(val))
    pattern = opts["PATTERN"]
    if opts["python_regex"]:
        return re.search(r"%s" % pattern, str(val))
    elif opts["word_regexp"]:
        if opts["ignore_case"]:
            return str(pattern).upper() == str(val).upper()
        else:
            return str(pattern) == str(val)
    elif not opts["word_regexp"]:
        if not opts["ignore_case"]:
            return str(pattern) in str(val)
        else:
            return str(pattern).upper() in str(val).upper()
    return None


def count_matching_strings(opts, cell, STRcount):
    if opts.get("compiled_regex"):
        for x in opts["compiled_regex"].findall(str(cell)):
            STRcount[0] += 1
    else:
        pattern = opts["PATTERN"]
        re_escaped = re.escape(str(pattern).upper())
        str_cell = str(cell).upper()
        if not opts["python_regex"]:
            for x in re.findall(re_escaped, str_cell):
                STRcount[0] += 1
        else:
            for x in re.findall(str(pattern), str(cell)):
                STRcount[0] += 1


def format_line_ending(opts):
    return "" if opts["null"] else "\n"


def format_single_match(opts, file, active_sheet, value):
    endswith = format_line_ending(opts)
    if opts["count"] or opts["files_with_match"] or opts["files_without_match"]:
        return ""

    if opts["with_filename"] and opts["with_sheetname"]:
        return f"{file}: {active_sheet}: {value}{endswith}"
    elif opts["with_filename"]:
        return f"{file}: {value}{endswith}"
    elif opts["with_sheetname"]:
        return f"{active_sheet}: {value}{endswith}"
    else:
        return f"{value}{endswith}"


def format_filename_and_sheetname(opts, file, active_sheet, linesArray):
    endswith = format_line_ending(opts)
    sep = str(opts["separator"])

    if opts["count"] or opts["files_with_match"] or opts["files_without_match"]:
        return ""

    if opts["with_filename"]:
        if opts["with_sheetname"]:
            return (
                file
                + ": "
                + active_sheet
                + ": "
                + sep
                + sep.join(map(str, linesArray))
                + endswith
            )
        else:
            return (
                file
                + ": "
                + sep
                + sep.join(map(str, linesArray))
                + endswith
            )
    else:
        if opts["with_sheetname"]:
            return (
                active_sheet
                + ": "
                + sep
                + sep.join(map(str, linesArray))
                + endswith
            )
        else:
            return sep.join(map(str, linesArray)) + endswith


def split_number_format(number_format):
    sections = []
    section = []
    in_quotes = False
    escaped = False
    for character in number_format:
        if escaped:
            section.append(character)
            escaped = False
        elif character == "\\":
            section.append(character)
            escaped = True
        elif character == '"':
            section.append(character)
            in_quotes = not in_quotes
        elif character == ";" and not in_quotes:
            sections.append("".join(section))
            section = []
        else:
            section.append(character)
    sections.append("".join(section))
    return sections


def excel_format_literal(value):
    literal = []
    in_quotes = False
    index = 0
    while index < len(value):
        character = value[index]
        if character == '"':
            in_quotes = not in_quotes
        elif character == "\\" and index + 1 < len(value):
            index += 1
            literal.append(value[index])
        elif character in "_*" and index + 1 < len(value):
            index += 1
        elif character != "_" or in_quotes:
            literal.append(character)
        index += 1
    return "".join(literal)


def format_excel_number(value, number_format):
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return value

    sections = split_number_format(number_format)
    if value < 0 and len(sections) > 1:
        section = sections[1]
        numeric_value = -value
    elif value == 0 and len(sections) > 2:
        section = sections[2]
        numeric_value = value
    else:
        section = sections[0]
        numeric_value = value

    section = re.sub(r"\[[^]]*\]", "", section)
    match = re.search(r"[0#?][0#?,]*(?:\.[0#?]+)?", section)
    if not match:
        return value

    numeric_pattern = match.group()
    integer_pattern, _, decimal_pattern = numeric_pattern.partition(".")
    decimals = len(decimal_pattern)
    grouping = "," in integer_pattern
    formatted = f"{numeric_value:,.{decimals}f}" if grouping else f"{numeric_value:.{decimals}f}"
    if decimal_pattern and "0" not in decimal_pattern:
        formatted = formatted.rstrip("0").rstrip(".")

    prefix = excel_format_literal(section[: match.start()])
    suffix = excel_format_literal(section[match.end() :])
    return prefix + formatted + suffix


def get_xlsx_book_dict(file):
    book = p.get_book_dict(file_name=file, skip_hidden_row_and_column=False)
    workbook = load_workbook(file, read_only=True, data_only=True)
    for sheet_name, rows in book.items():
        worksheet = workbook[sheet_name]
        for row, cells in zip(rows, worksheet.iter_rows()):
            for column_index, (value, cell) in enumerate(zip(row, cells)):
                row[column_index] = format_excel_number(value, cell.number_format)
    workbook.close()
    return book


def get_xls_book_dict(file):
    workbook = xlrd.open_workbook(file, formatting_info=True)
    book = {}
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        rows = []
        for row_index in range(sheet.nrows):
            row = []
            for col_index in range(sheet.ncols):
                cell_type = sheet.cell_type(row_index, col_index)
                value = sheet.cell_value(row_index, col_index)
                if cell_type == xlrd.XL_CELL_NUMBER:
                    xf_index = sheet.cell_xf_index(row_index, col_index)
                    xf = workbook.xf_list[xf_index]
                    fmt_key = xf.format_key
                    fmt = workbook.format_map.get(fmt_key)
                    if fmt is not None and fmt.format_str:
                        value = format_excel_number(value, fmt.format_str)
                row.append(value)
            rows.append(row)
        book[sheet.name] = rows
    return book


def get_ods_book_dict(file):
    import zipfile
    import xml.etree.ElementTree as ET

    def get_attr(elem, name):
        for key, val in elem.attrib.items():
            if key == name or key.endswith("}" + name):
                return val
        return None

    def get_paragraph_text(p_node):
        parts = []

        def _walk(node):
            if node.text:
                parts.append(node.text)
            for child in node:
                tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                if tag == "s":
                    c = get_attr(child, "c") or "1"
                    try:
                        count = int(c)
                    except ValueError:
                        count = 1
                    parts.append(" " * count)
                elif tag == "tab":
                    parts.append("\t")
                elif tag == "line-break":
                    parts.append("\n")
                else:
                    _walk(child)
                if child.tail:
                    parts.append(child.tail)

        _walk(p_node)
        return "".join(parts)

    def extract_cell_text(cell, ns):
        paragraphs = [get_paragraph_text(p) for p in cell.findall(".//text:p", ns)]
        return "\n".join(paragraphs) if paragraphs else ""

    try:
        with zipfile.ZipFile(file) as zf:
            root = ET.fromstring(zf.read("content.xml"))

        ns = {
            "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
            "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
            "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
            "style": "urn:oasis:names:tc:opendocument:xmlns:style:1.0",
        }

        book = {}
        for table in root.findall(".//table:table", ns):
            sheet_name = get_attr(table, "name") or "Sheet"
            rows = []
            for row in table.findall("table:table-row", ns):
                row_repeated_str = get_attr(row, "number-rows-repeated")
                row_repeated = int(row_repeated_str) if row_repeated_str else 1
                values = []
                for cell in row:
                    tag = cell.tag.split("}")[-1] if "}" in cell.tag else cell.tag
                    if tag not in ("table-cell", "covered-table-cell"):
                        continue
                    col_repeated_str = get_attr(cell, "number-columns-repeated")
                    col_repeated = int(col_repeated_str) if col_repeated_str else 1
                    if tag == "covered-table-cell":
                        text = ""
                    else:
                        text = extract_cell_text(cell, ns)
                        if not text:
                            value_type = get_attr(cell, "value-type")
                            if value_type in ("float", "currency", "percentage"):
                                val = get_attr(cell, "value")
                                if val is not None:
                                    try:
                                        num = float(val)
                                        text = int(num) if num.is_integer() else num
                                    except ValueError:
                                        text = val
                                else:
                                    text = ""
                            elif value_type == "boolean":
                                val = get_attr(cell, "boolean-value")
                                text = val.lower() == "true" if val else ""
                            elif value_type == "date":
                                text = get_attr(cell, "date-value") or ""
                            elif value_type == "time":
                                text = get_attr(cell, "time-value") or ""
                            else:
                                text = ""
                    values.extend([text] * col_repeated)

                while values and (values[-1] == "" or values[-1] is None):
                    values.pop()

                for _ in range(row_repeated):
                    rows.append(list(values))

            while rows and not any(c != "" and c is not None for c in rows[-1]):
                rows.pop()

            book[sheet_name] = rows
        return book
    except Exception:
        return p.get_book_dict(file_name=file)


def process_single_file(file, opts):
    stdout_lines = []
    stderr_lines = []
    SumOfROW, SumOfCELL, SumOfSTR = [0], [0], [0]

    if not opts["debug"]:
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            message="Unknown extension is not supported and will be removed",
        )
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            message="Data Validation extension is not supported and will be removed",
        )
        warnings.filterwarnings(
            "ignore",
            category=DeprecationWarning,
            message=".*Flags not at the start of the expression*.",
        )
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            message="Conditional Formatting extension is not supported and will be removed",
        )
        warnings.filterwarnings(
            "ignore",
            category=UserWarning,
            message="Cannot parse header or footer so it will be ignored",
        )
        logging.disable(logging.WARNING)
    else:
        warnings.resetwarnings()
        stdout_lines.append(f"-- debug mode: {file}\n")

    try:
        if file.endswith((".xlsx", ".XLSX", ".xlsm", ".XLSM")):
            book = get_xlsx_book_dict(file)
        elif file.endswith((".xls", ".XLS")):
            book = get_xls_book_dict(file)
        elif file.endswith((".ods", ".ODS")):
            book = get_ods_book_dict(file)
        else:
            book = p.get_book_dict(file_name=file)
    except KeyboardInterrupt:
        raise
    except Exception:
        stderr_lines.append(
            f"Error:\tUnsupported format, password protected or corrupted file: {file}\n"
        )
        return {
            "file": file,
            "stdout": stdout_lines,
            "stderr": stderr_lines,
            "counts": (0, 0, 0),
        }

    endswith = format_line_ending(opts)

    if opts["files_with_match"]:
        if check_optional_args(opts, book):
            stdout_lines.append(f"{file}{endswith}")
        return {
            "file": file,
            "stdout": stdout_lines,
            "stderr": stderr_lines,
            "counts": (0, 0, 0),
        }
    elif opts["files_without_match"]:
        if not check_optional_args(opts, book):
            stdout_lines.append(f"{file}{endswith}")
        return {
            "file": file,
            "stdout": stdout_lines,
            "stderr": stderr_lines,
            "counts": (0, 0, 0),
        }

    if opts["column"]:
        COLcount, CELLcount, STRcount = [0], [0], [0]
        for key, item in book.items():
            if not item:
                continue
            max_cols = max(len(row) for row in item)
            for col_idx in range(max_cols):
                column = [
                    row[col_idx] if col_idx < len(row) else ""
                    for row in item
                ]
                col_has_match = any(
                    check_optional_args(opts, cell) for cell in column
                )
                if not col_has_match:
                    continue

                COLcount[0] += 1
                if opts["count"]:
                    for cell in column:
                        if check_optional_args(opts, cell):
                            CELLcount[0] += 1
                            count_matching_strings(opts, cell, STRcount)
                else:
                    for cell in column:
                        out = format_single_match(opts, file, key, cell)
                        if out:
                            stdout_lines.append(out)

        if opts["count"] and COLcount[0] > 0:
            if opts["with_sheetname"] or opts["with_filename"]:
                stdout_lines.append(
                    f"{file} : {COLcount[0]} Columns,  {CELLcount[0]} Cells,  {STRcount[0]} Strings{endswith}"
                )
            SumOfROW[0] = COLcount[0]
            SumOfCELL[0] = CELLcount[0]
            SumOfSTR[0] = STRcount[0]
    else:
        ROWcount, CELLcount, STRcount = [0], [0], [0]
        for key, item in book.items():
            for line in item:
                AuxFlag = False
                for cell in line:
                    if check_optional_args(opts, cell):
                        if opts["count"]:
                            AuxFlag = True
                            CELLcount[0] += 1
                            count_matching_strings(opts, cell, STRcount)
                        else:
                            AuxFlag = True
                            ROWcount[0] -= 1

                if AuxFlag:
                    ROWcount[0] += 1
                    out = format_filename_and_sheetname(opts, file, key, line)
                    if out:
                        stdout_lines.append(out)

        if opts["count"] and ROWcount[0] > 0:
            if opts["with_sheetname"] or opts["with_filename"]:
                stdout_lines.append(
                    f"{file} : {ROWcount[0]} Rows,  {CELLcount[0]} Cells,  {STRcount[0]} Strings{endswith}"
                )
            SumOfROW[0] = ROWcount[0]
            SumOfCELL[0] = CELLcount[0]
            SumOfSTR[0] = STRcount[0]

    return {
        "file": file,
        "stdout": stdout_lines,
        "stderr": stderr_lines,
        "counts": (SumOfROW[0], SumOfCELL[0], SumOfSTR[0]),
    }


def SEARCH(File_List, opts):
    SumOfROW, SumOfCELL, SumOfSTR = [], [], []
    jobs = opts.get("jobs", 1)
    if jobs <= 0:
        jobs = os.cpu_count() or 1

    def process_result(res):
        try:
            for err in res["stderr"]:
                sys.stderr.write(err)
                sys.stderr.flush()
            for out in res["stdout"]:
                sys.stdout.write(out)
                sys.stdout.flush()
        except BrokenPipeError:
            try:
                sys.stderr.close()
            except Exception:
                pass
            sys.exit(0)
        r, c, s = res["counts"]
        if opts["count"] and (r > 0 or c > 0 or s > 0):
            SumOfROW.append(r)
            SumOfCELL.append(c)
            SumOfSTR.append(s)

    if jobs > 1 and len(File_List) > 1:
        max_workers = jobs
        worker_fn = functools.partial(process_single_file, opts=opts)
        try:
            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                results = executor.map(worker_fn, File_List)
                for res in results:
                    process_result(res)
        except KeyboardInterrupt:
            sys.exit(0)
    else:
        for file in File_List:
            try:
                res = process_single_file(file, opts)
                process_result(res)
            except KeyboardInterrupt:
                sys.exit(0)

    if opts["count"]:
        if not (opts["files_with_match"] or opts["files_without_match"]):
            GROUPS, CELLS, STRINGS = sum(SumOfROW), sum(SumOfCELL), sum(SumOfSTR)
            group_label = "Columns" if opts["column"] else "Rows"
            print(
                "Search results: ",
                GROUPS,
                group_label + ", ",
                CELLS,
                "Cells, ",
                STRINGS,
                "Strings",
            )


def main():

    help_text = """positional arguments:
  PATTERN                    use PATTERN as the pattern to search for.
  FILE                       file or path to folder

options:
  -h, --help                 show this help message and exit.
  -V, --version              display version information and exit.
  -P, --python-regex         PATTERN is a Python regular expression.
  -E, --extended-regexp      PATTERN is a POSIX extended regular expression (Default).
  -F, --fixed-strings        interpret PATTERN as fixed strings, not regular expressions.
  -i, --ignore-case          ignore case distinctions.
  -w, --word-regexp          force PATTERN to match only whole words.
  -c, --count                print only a count of matches per file.
  -r, --recursive            search directories recursively.
  -H, --with-filename        print the file name for each match.
  -N, --with-sheetname       print the sheet name for each match.
  -l, --files-with-match     print only names of FILEs with match pattern.
  -L, --files-without-match  print only names of FILEs with no match pattern.
  -S, --separator SEPARATOR  define custom list separator for output, the default is TAB.
  -j, --jobs JOBS            number of CPU cores/processes to use for search (default: 1).
      --row                  search rows and print matching rows (default).
      --column               search columns and print whole matching columns vertically.

examples:
    xlsxgrep -i "foo" foobar.xlsx
    xlsxgrep -c -H "(?i)foo|bar" /folder

For more details refer to the man page.
"""
    parser = argparse.ArgumentParser(
        add_help=False,  # epilog=example_text,
        description=dedent(help_text),
        prog="xlsxgrep",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        usage=dedent(
            """
	    xlsxgrep [-h] [-V] [-P] [-E] [-F] [-i] [-w] [-c] [-r] [-H] [-N] [-l] [-L] [-S SEPARATOR] 
                [-Z] [-j JOBS] [--row | --column] [-d] PATTTERN FILE [FILE ...]


            """
        ).strip(),
    )
    parser.add_argument(
        "-h", "--help", action="help", help=argparse.SUPPRESS
    )
    parser.add_argument(
        "PATTERN", help=argparse.SUPPRESS, type=str
    )
    parser.add_argument(
        "-V",
        "--version",
        help=argparse.SUPPRESS,
        # help="display version information and exit.",
        action="version",
        version=dedent("\n".join(VERSION_INFO) + "\n"),
    )
    parser.add_argument(
        "FILE", help=argparse.SUPPRESS, nargs="+", action="append",
    )
    parser.add_argument(
        "-P",
        "--python-regex",
        help=argparse.SUPPRESS,
        # help="PATTERN is a Python regular expression. This is the default.",
        required=False,
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "-E",
        "--extended-regexp",
        "--extended-regex",
        help=argparse.SUPPRESS,
        required=False,
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "-F",
        "--fixed-strings",
        help=argparse.SUPPRESS,
        # help="interpret PATTERN as fixed strings, not regular expressions.",
        required=False,
        action="store_true",
        default=False,
    )
    parser.add_argument(
        "-i",
        "--ignore-case",
        help=argparse.SUPPRESS,
        # help="ignore case distinctions.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-w",
        "--word-regexp",
        help=argparse.SUPPRESS,
        # help="force PATTERN to match only whole words.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-c",
        "--count",
        help=argparse.SUPPRESS,
        # help="print only a count of matches per file.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        help=argparse.SUPPRESS,
        # help="search directories recursively.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-H",
        "--with-filename",
        help=argparse.SUPPRESS,
        # help="print the file name for each match.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-N",
        "--with-sheetname",
        help=argparse.SUPPRESS,
        # help="print the sheet name for each match.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-l",
        "--files-with-match",
        help=argparse.SUPPRESS,
        # help="print only names of FILEs with match pattern.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-L",
        "--files-without-match",
        help=argparse.SUPPRESS,
        # help="print only names of FILEs with no match pattern.",
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-S",
        "--separator",
        # help="define custom list separator for output, the default is TAB.",
        help=argparse.SUPPRESS,
        required=False,
        default="\t",
        type=str,
    )
    parser.add_argument(
        "-Z",
        "--null",
        help=argparse.SUPPRESS,
        required=False,
        action="store_true",
    )
    parser.add_argument(
        "-j",
        "--jobs",
        help=argparse.SUPPRESS,
        required=False,
        default=1,
        type=int,
    )
    parser.add_argument(
        "-d",
        "--debug",
        help=argparse.SUPPRESS,
        required=False,
        default=False,
        action="store_true",
    )
    search_mode = parser.add_mutually_exclusive_group()
    search_mode.add_argument(
        "--row",
        help=argparse.SUPPRESS,
        action="store_true",
    )
    search_mode.add_argument(
        "--column",
        help=argparse.SUPPRESS,
        action="store_true",
    )

    if len(sys.argv) == 1:
        parser.print_usage(sys.stderr)
        print("Type 'xlsxgrep --help' for more information.")
        sys.exit(1)

    args = parser.parse_args()

    def ActivateDebug():
        if args.debug == False:
            # Some debug options
            # - Supress unsupported file extensions warnings.
            # 'UserWarning: Data Validation extension is not supported and will be removed'. (module=openpyxl)
            # 'UserWarning: Unknown extension is not supported and will be removed'.         (module=openpyxl)
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                message="Unknown extension is not supported and will be removed",
            )
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                message="Data Validation extension is not supported and will be removed",
            )
            # - Ignore deprecated python regex warnings.
            # 'DeprecationWarning: Flags not at the start of the expression 'foo|(?i)bar'.   (module=re)
            warnings.filterwarnings(
                "ignore",
                category=DeprecationWarning,
                message=".*Flags not at the start of the expression*.",
            )
            # - Supress Conditional Formatting extension not supported and Cannot parse header or footer warning.
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                message="Conditional Formatting extension is not supported and will be removed",
            )
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                message="Cannot parse header or footer so it will be ignored",
            )
            # - Disable all warnings in openpyxl
            warnings.filterwarnings(
                "ignore", category=UserWarning, module="openpyxl")
            # - Disable all logging warnings
            logging.disable(logging.WARNING)
        else:
            print("--version info: "+" ".join(VERSION_INFO))

            pass

    ActivateDebug()

    # Valid Python / POSIX Regex Check

    def Check_Regex():
        compiled = None

        if args.python_regex:
            if args.extended_regexp:
                sys.exit(
                    "xlsxgrep: --python-regex cannot be used together with: -E"
                )
            if args.fixed_strings:
                sys.exit(
                    "xlsxgrep: --python-regex cannot be used together with: -F"
                )
            try:
                args.python_regex = True
                args.posix_ere = False
                compiled = re.compile(args.PATTERN)
            except re.error:
                sys.exit(
                    "Error:  Not valid Python Regular Expression. For fixed strings use flag: -F"
                )
            return compiled

        if args.fixed_strings:
            if args.extended_regexp:
                sys.exit(
                    "xlsxgrep: --fixed-strings cannot be used together with: -E"
                )
            args.python_regex = False
            args.posix_ere = False
            return compiled

        # Default behavior is POSIX ERE unless Python regex or fixed strings are explicitly selected.
        args.python_regex = False
        args.posix_ere = True
        pattern = translate_posix_ere(args.PATTERN)
        if args.word_regexp:
            pattern = r"\b(?:" + pattern + r")\b"
        flags = re.IGNORECASE if args.ignore_case else 0
        try:
            compiled = re.compile(pattern, flags)
        except re.error:
            sys.exit(
                "Error:  Not valid POSIX Extended Regular Expression."
            )

        return compiled

    compiled_regex = Check_Regex()

    # Checking file or folder format and destination

    def File_And_Path_Location():
        File_List = []
        fileTypes = (
            ".xls",
            ".XLS",
            ".xlsx",
            ".XLSX",
            ".ods",
            ".ODS",
            ".csv",
            ".CSV",
            ".tsv",
            ".TSV",
            ".xlsm",
            ".XLSM",
        )
        for i in args.FILE[0]:

            if (Path(i).is_file() is False) and (Path(i).is_dir() is False):
                exit(str(i) + " File or folder not found. ")

            elif Path(i).is_file() and str(Path(i)).endswith(fileTypes):
                File_List.append(str(Path(i)))

            elif Path(i).is_dir():
                if args.recursive == True:
                    for child in Path(i).rglob("*"):
                        if str(child).endswith(fileTypes):
                            File_List.append(str(child))
                else:
                    for child in Path(i).iterdir():
                        if str(child).endswith(fileTypes):
                            File_List.append(str(child))

            elif (Path(i).is_file() and str(Path(i)).endswith(fileTypes)) == False:
                # perform file check
                print("Error:   Unsupported file format: ",
                      Path(i), file=sys.stderr)

        opts = {
            "PATTERN": args.PATTERN,
            "python_regex": args.python_regex,
            "posix_ere": getattr(args, "posix_ere", False),
            "extended_regexp": args.extended_regexp,
            "fixed_strings": args.fixed_strings,
            "ignore_case": args.ignore_case,
            "word_regexp": args.word_regexp,
            "compiled_regex": compiled_regex,
            "count": args.count,
            "recursive": args.recursive,
            "with_filename": args.with_filename,
            "with_sheetname": args.with_sheetname,
            "files_with_match": args.files_with_match,
            "files_without_match": args.files_without_match,
            "separator": args.separator,
            "null": args.null,
            "debug": args.debug,
            "row": args.row,
            "column": args.column,
            "jobs": args.jobs,
        }

        SEARCH(File_List, opts)

    File_And_Path_Location()


if __name__ == "__main__":
    main()
